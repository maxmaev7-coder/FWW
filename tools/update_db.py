import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WB_PATH = ROOT / "All cards wave 1-9 21_02_2024 with categories for builder (2).xlsx"
UNITS_JSON = ROOT / "db" / "units.json"
ITEMS_JSON = ROOT / "db" / "items.json"

FACTION_COLUMNS_UNITS = [
    "B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V"
]
WEAPON_COLUMNS_UNITS = {
    "W": "Melee",
    "X": "Pistol",
    "Y": "Rifle",
    "Z": "Heavy Weapon",
    "AA": "Grenade",
    "AB": "Power Armor",
}
ACCESS_COLUMNS_UNITS = {
    "AC": "Upgrades",
    "AD": "Wasteland Items",
    "AE": "Advanced Items",
    "AF": "High Tech Items",
    "AG": "Usable Items",
    "AH": "Robots Items",
    "AI": "Automatron Items",
    "AJ": "Creature Items",
    "AK": "Dog Items",
    "AL": "Super Mutant Items",
    "AM": "Standart Item",
    "AN": "Faction Items",
}

FACTION_COLUMNS_ITEMS = [
    "D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S"
]
WEAPON_COLUMNS_ITEMS = {
    "T": "Melee",
    "U": "Pistol",
    "V": "Rifle",
    "W": "Heavy Weapon",
    "X": "Grenade",
    "Y": "Mines",
}
CATEGORY_COLUMNS_ITEMS = {
    "Z": "Chem",
    "AA": "Alcohol",
    "AB": "Food",
    "AC": "Armor",
    "AD": "Clothes",
    "AE": "Gear",
    "AF": "Mod",
    "AG": "Perks",
    "AH": "Leader",
    "AI": "Power Armor",
    "AJ": "Upgrades",
    "AK": "Wasteland Items",
    "AL": "Advanced Items",
    "AM": "High Tech Items",
    "AN": "Usable Items",
    "AO": "Robots Items",
    "AP": "Automatron Items",
    "AQ": "Creature Items",
    "AR": "Dog Items",
    "AS": "Super Mutant Items",
    "AT": "Standart Item",
}
MOD_TARGET_COLUMNS = {
    "BI": "Melee",
    "BJ": "Pistol",
    "BK": "Rifle",
    "BL": "Heavy Weapon",
    "BM": "Armor",
    "BN": "Power Armor",
    "BO": "Robot",
    "BP": "Animal",
}
FACTION_LIMIT_COLUMNS = {
    "CS": "BoS",
    "CT": "Caesar's Legion",
    "CU": "Creatures",
    "CV": "Enclave",
    "CW": "Institute",
    "CX": "NCR",
    "CY": "Raiders",
    "CZ": "Robots",
    "DA": "Super Mutants",
    "DB": "Survivors",
    "DC": "CoA",
    "DD": "Forged",
    "DE": "Gunners",
    "DF": "Railroad",
    "DG": "The Slog",
}


def slugify(name: str) -> str:
    base = name.lower()
    base = re.sub(r"[^a-z0-9]+", "-", base)
    base = re.sub(r"-+", "-", base).strip("-")
    return base or "item"


def normalize(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", text.lower())


def parse_equipped(raw: str, item_lookup: dict[str, str]) -> list[str]:
    if not raw:
        return []
    raw = raw.replace("\n", ",")
    # remove long descriptive fragments that definitely are not card names
    pieces = re.split(r",| and | & ", raw)
    names: list[str] = []
    pending_mult = 1
    for piece in pieces:
        piece = piece.strip()
        if not piece:
            continue
        m = re.match(r"x(\d+)\s*(.*)", piece, re.IGNORECASE)
        if m:
            pending_mult = int(m.group(1))
            piece = m.group(2).strip()
        if not piece:
            continue
        key = normalize(piece)
        if not key:
            pending_mult = 1
            continue
        match_id = None
        for norm_name, item_id in item_lookup.items():
            if norm_name == key:
                match_id = item_id
                break
        if not match_id:
            # try partial match
            for norm_name, item_id in item_lookup.items():
                if key in norm_name:
                    match_id = item_id
                    break
        if match_id:
            names.extend([match_id] * max(1, pending_mult))
        pending_mult = 1
    return names


def load_items() -> list[dict]:
    if not ITEMS_JSON.exists():
        return []
    with ITEMS_JSON.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def load_units() -> list[dict]:
    if not UNITS_JSON.exists():
        return []
    with UNITS_JSON.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def main() -> None:
    wb = load_workbook(WB_PATH, data_only=True)
    items_sheet = wb["Items (2)"]
    units_sheet = wb["Units"]

    old_items = load_items()
    items_by_name = {item["name"]: item for item in old_items}
    name_to_id = {normalize(item["name"]): item["id"] for item in old_items}

    new_items: list[dict] = []
    for row in range(2, items_sheet.max_row + 1):
        name = items_sheet[f"A{row}"].value
        if not name:
            continue
        name = str(name).strip()
        if not name:
            continue
        base = items_by_name.get(name, {"id": slugify(name)})
        item_id = base.get("id") or slugify(name)
        primary = items_sheet[f"B{row}"].value
        category = items_sheet[f"C{row}"].value
        factions = []
        for col in FACTION_COLUMNS_ITEMS:
            val = items_sheet[f"{col}{row}"].value
            if isinstance(val, str) and val.strip().lower() == "x":
                header = items_sheet[f"{col}1"].value or ""
                if header:
                    factions.append(str(header))
        weapon = {key: False for key in WEAPON_COLUMNS_ITEMS.values()}
        for col, label in WEAPON_COLUMNS_ITEMS.items():
            val = items_sheet[f"{col}{row}"].value
            weapon[label] = isinstance(val, str) and val.strip().lower() == "x"
        cats = {key: False for key in CATEGORY_COLUMNS_ITEMS.values()}
        for col, label in CATEGORY_COLUMNS_ITEMS.items():
            val = items_sheet[f"{col}{row}"].value
            cats[label] = isinstance(val, str) and val.strip().lower() == "x"
        is_mod = (str(primary).strip().lower() == "mod") or cats.get("Mod", False)
        mod_targets: list[str] = []
        for col, label in MOD_TARGET_COLUMNS.items():
            val = items_sheet[f"{col}{row}"].value
            if isinstance(val, str) and val.strip().lower() == "x":
                mod_targets.append(label)
        unique = False
        special_flag = items_sheet[f"BA{row}"].value
        if isinstance(special_flag, str) and special_flag.strip().lower() == "x":
            unique = True
        cost_candidates = [
            items_sheet[f"BG{row}"].value,
            items_sheet[f"BF{row}"].value,
            items_sheet[f"BD{row}"].value,
            items_sheet[f"BE{row}"].value,
        ]
        cost = next((c for c in cost_candidates if isinstance(c, (int, float))), base.get("cost"))
        faction_limits = {}
        for col, label in FACTION_LIMIT_COLUMNS.items():
            raw = items_sheet[f"{col}{row}"].value
            if raw is None:
                continue
            try:
                num = int(str(raw).strip())
            except ValueError:
                continue
            faction_limits[label] = num
        item_record = {
            "id": item_id,
            "name": name,
            "primary": primary or "",
            "category": category or "",
            "cost": int(cost) if isinstance(cost, (int, float)) else 0,
            "factions": factions,
            "weapon": weapon,
            "cats": cats,
            "is_mod": bool(is_mod),
            "mod_targets": mod_targets,
            "unique": bool(unique),
            "faction_limits": faction_limits,
        }
        new_items.append(item_record)

    new_items.sort(key=lambda x: (x["primary"], x["name"]))

    with ITEMS_JSON.open("w", encoding="utf-8") as fh:
        json.dump(new_items, fh, ensure_ascii=False, indent=2)

    old_units = load_units()
    units_by_name = {unit["name"]: unit for unit in old_units}

    item_lookup = name_to_id

    new_units: list[dict] = []
    for row in range(3, units_sheet.max_row + 1):
        name = units_sheet[f"A{row}"].value
        if not name:
            continue
        name = str(name).strip()
        if not name:
            continue
        base = units_by_name.get(name, {"id": slugify(name)})
        unit_id = base.get("id") or slugify(name)
        factions = []
        for col in FACTION_COLUMNS_UNITS:
            val = units_sheet[f"{col}{row}"].value
            if isinstance(val, str) and val.strip().lower() == "x":
                header = units_sheet[f"{col}2"].value or ""
                if header:
                    factions.append(str(header))
        prereq = base.get("prereq", {})
        for col, label in WEAPON_COLUMNS_UNITS.items():
            val = units_sheet[f"{col}{row}"].value
            prereq[label] = isinstance(val, str) and val.strip().lower() == "x"
        access = base.get("access", {})
        for col, label in ACCESS_COLUMNS_UNITS.items():
            val = units_sheet[f"{col}{row}"].value
            access[label] = isinstance(val, str) and val.strip().lower() == "x"
        cost_candidates = [
            units_sheet[f"AT{row}"].value,
            units_sheet[f"AU{row}"].value,
            units_sheet[f"AS{row}"].value,
            units_sheet[f"AQ{row}"].value,
        ]
        cost = next((c for c in cost_candidates if isinstance(c, (int, float))), base.get("cost"))
        unique_flag = units_sheet[f"AX{row}"].value
        unique = isinstance(unique_flag, str) and unique_flag.strip().lower() == "x"
        raw_equipped = units_sheet[f"AO{row}"].value
        equipped = base.get("equipped", [])
        parsed = parse_equipped(str(raw_equipped) if raw_equipped else "", item_lookup)
        if parsed:
            equipped = parsed
        unit_record = {
            "id": unit_id,
            "name": name,
            "factions": factions,
            "prereq": prereq,
            "access": access,
            "unique": bool(unique),
            "cost": int(cost) if isinstance(cost, (int, float)) else base.get("cost", 0),
            "equipped": equipped,
        }
        new_units.append(unit_record)

    new_units.sort(key=lambda x: x["name"])

    with UNITS_JSON.open("w", encoding="utf-8") as fh:
        json.dump(new_units, fh, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
